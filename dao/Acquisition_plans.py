import pandas as pd
import requests
import arrow
import openpyxl
from openpyxl import load_workbook

table = "5997216"
path = 'C:/Users/xiaofei.yu/Desktop\采集预案小程序'
month = arrow.now().shift(months = -1).strftime('%Y%m')
month1 = arrow.now().strftime('%Y%m')
monthly = [month,month1]
day = arrow.now().strftime('%Y%m%d')
# 获取token
def get_token():
    result = requests.post('http://services.kurite.cn/huoban/getToken',
                           data={'id': '12067898212', 'pwd': 'xjoijsdif2112'})
    return result.json()


# 根据当前token，获取对应表格的信息
def get_table(token):
    global data
    param = {'token': token, 'table': table}  # , 'name': ''
    try:
        result = requests.get('http://services.kurite.cn/huoban/getFields', params=param)
        if result.content:
            data = result.json()
        else:
            data = '暂未获取到数据'
    except Exception as e:
      print('获取token失败')
    return data


# 获取表格内容
def get_table_info(current_token):
    # 请求体必要信息
    detail_datas = []
    datas = {
        'token': current_token,
        'table': table
        , 'offset': 0
    }
    # 调用post请求，获取response
    detail_info = requests.post('http://services.kurite.cn/huoban/getTableInfo', data=datas)
    row_total = detail_info.json()['data']['total']
    offset_num = int(row_total / 500) + 1
    for offset in range(offset_num):
        datas = {
            'token': current_token,
            'table': table
            , 'offset': (offset * 500)
        }
        # 调用post请求，获取response
        detail_info = requests.post('http://services.kurite.cn/huoban/getTableInfo', data=datas)
        detail_datas.append(detail_info)
    return detail_datas


def create_ctr_table():
    res_json = get_token()
    # 获取response中的token
    current_token = res_json['data']['token']
    # 获取表格内容
    table_info = get_table_info(current_token)
    tb = pd.DataFrame()
    for i in range(len(table_info)):
        tab_json = table_info[i].json()
        zkb_data = tab_json['data']['tableInfo']
        df = pd.DataFrame(zkb_data)
        tb = pd.concat([tb, df])
    print('总控表导出成功')
    # tb.to_csv('111111111111.csv')
    return tb


def mapping(df_mapping, code):
    df_mapping['符号'] = df_mapping['符号'].astype(str)
    flag = str(code)[:2]
    if flag in df_mapping['符号'].tolist():
        df_mapping = df_mapping[df_mapping['符号'] == flag]
        line = df_mapping['产线'].tolist()[0]
        return line
    else:
        return '其他'

def write_excel(dataframe, excelWriter, sheetname):
    """
    数据写入到Excel,可以写入不同的sheet
    """
    excelWriter.book = load_workbook(excelWriter.path)
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sheetname, index=None)
    excelWriter.close()

def disosal_Acquisition():
    df_mapping = pd.read_excel(path+'/产线对应关系.xlsx')
    df = create_ctr_table()
    df = df[df['q收集标记'].apply(lambda a  : str(a) in monthly)]
    df = df[['q收集标记','经销商代码','经销商名称','q收集方式','q级别','q手工标记']]
    df.rename(columns = {'q收集标记':'年月'},inplace=True)
    df['产线'] = df.apply(lambda x : mapping(df_mapping,x['经销商代码']),axis=1)
    df = df[['年月','产线','经销商代码','经销商名称','q收集方式','q级别','q手工标记']]
    df = df[df['产线'] != '其他']
    df_data = pd.read_excel(path+'/返回情况表'+day+'.xlsx')
    df_data = df_data[df_data['q收集方式'].apply(lambda a :str(a) in ['ADI','FTP'])]
    df_data = df_data[df_data['进销存类型'].apply(lambda a :str(a) in ['xiao','nan'])]
    df_data = df_data[df_data['状态'].apply(lambda a: str(a) == '未返回')]
    df_data_list = df_data['经销商代码'].tolist()
    for i,row in df.iterrows():
        if row['经销商代码'] in df_data_list:
            row['q手工标记'] = str(month)

    wb = openpyxl.load_workbook(path+'/采集预案模板.xlsx')
    sheet1 = wb.active
    df_dmk = df[df['产线'].apply(lambda a: str(a) == str('DMK'))]
    y_count = df_dmk.shape[0]
    y_adi_count = df_dmk[df_dmk['q收集方式'].apply(lambda a: str(a) in ['ADI', 'FTP', ])].shape[0]
    y_mnl_count = df_dmk[df_dmk['q收集方式'].apply(lambda a: str(a) not in ['ADI', 'FTP', ])].shape[0]
    df_dmk = df_dmk[df_dmk['q收集方式'].apply(lambda a: str(a) in ['ADI', 'FTP', ])]
    y_adi_mnl_count = df_dmk[df_dmk['q手工标记'].apply(lambda a: str(a) != 'nan')].shape[0]
    sheet1.cell(2, 2, y_count)
    sheet1.cell(3, 2, y_adi_count)
    sheet1.cell(4, 2, y_mnl_count)
    sheet1.cell(5, 2, y_adi_mnl_count)
    df_leo = df[df['产线'].apply(lambda a: str(a) == str('LEO'))]
    l_count = df_leo.shape[0]
    l_adi_count = df_leo[df_leo['q收集方式'].apply(lambda a: str(a) in ['ADI', 'FTP', ])].shape[0]
    l_mnl_count = df_leo[df_leo['q收集方式'].apply(lambda a: str(a) not in ['ADI', 'FTP', ])].shape[0]
    df_leo = df_leo[df_leo['q收集方式'].apply(lambda a: str(a) in ['ADI', 'FTP', ])]
    l_adi_mnl_count = df_leo[df_leo['q手工标记'].apply(lambda a: str(a) != 'nan')].shape[0]
    sheet1.cell(2, 3, l_count)
    sheet1.cell(3, 3, l_adi_count)
    sheet1.cell(4, 3, l_mnl_count)
    sheet1.cell(5, 3, l_adi_mnl_count)
    sheet1.cell(2, 4, l_count+y_count)
    sheet1.cell(3, 4, l_adi_count+y_adi_count)
    sheet1.cell(4, 4, l_mnl_count+y_mnl_count)
    sheet1.cell(5, 4, l_adi_mnl_count+y_adi_mnl_count)

    wb.save(path + '/采集预案' + day + '.xlsx')
    excelWriter = pd.ExcelWriter(path + '/采集预案' + day + '.xlsx', engine='openpyxl')
    write_excel(df, excelWriter, '详情')



if __name__ == '__main__':
    disosal_Acquisition()





